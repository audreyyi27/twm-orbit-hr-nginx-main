from datetime import datetime, timedelta
from io import BytesIO
from typing import Literal, Optional
from fastapi import APIRouter, Depends,HTTPException, Query
import pytz
from .. import schemas, models
from ..db import get_db
from ..deps import get_current_user, get_current_user_hr, parse_new_candidate
from fastapi.responses import Response
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, LongTable, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import and_, asc, desc, func, insert, or_, select, update, cast, String
from sqlalchemy.orm import load_only
router = APIRouter(prefix="/reports", tags=['reports'])

LOCAL_TZ = pytz.timezone('Asia/Jakarta')  # Change to your timezone
END_OF_DAY_HOUR = 18  # 6 PM

@router.get(
    "",
    summary="Export candidate report (Excel or PDF)",
    description=(
        "Generates and returns a candidate report filtered by status and date range. "
        "Dates are interpreted in the local timezone (Asia/Jakarta). "
        "Set `file_type=excel` for an `.xlsx` download or `file_type=pdf` for a landscape A4 PDF."
    ),
    responses={
        200: {
            "description": "Report file stream",
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {
                    "schema": {"type": "string", "format": "binary"},
                    "examples": {
                        "excel": {
                            "summary": "Excel workbook",
                            "description": "XLSX file named `candidates.xlsx`",
                            "value": "<binary stream>",
                        }
                    },
                },
                "application/pdf": {
                    "schema": {"type": "string", "format": "binary"},
                    "examples": {
                        "pdf": {
                            "summary": "PDF document",
                            "description": "Landscape A4 PDF named `candidates.pdf`",
                            "value": "<binary stream>",
                        }
                    },
                },
            },
        },
        400: {
            "description": "Bad request (invalid dates) or no matching data for the given filters",
        },
        401: {"description": "Unauthorized"},
        403: {"description": "Forbidden (requires hr_admin)"},
    },
)
async def get_candidate_report(
    candidate_status: Optional[schemas.CandidateStatusEnum] = Query(
        None,
        description="Filter by candidate processed status. If omitted, all statuses are included.",
    ),
    start_date: Optional[str] = Query(
        None,
        description="Filter start date in YYYY-MM-DD (local time Asia/Jakarta). Inclusive.",
        examples=["2025-01-01"],
    ),
    end_date: Optional[str] = Query(
        None,
        description="Filter end date in YYYY-MM-DD (local time Asia/Jakarta). Inclusive.",
        examples=["2025-01-31"],
    ),
    file_type: Optional[Literal["pdf", "excel"]] = Query(
        "excel",
        description="Output format. `excel` returns .xlsx, `pdf` returns a PDF.",
    ),
    db: AsyncSession = Depends(get_db),
    current=Depends(get_current_user_hr),
):
  # Only select fields that exist in your actual DB (Indonesian field names)
  query = select(models.Candidates).options(load_only(
        models.Candidates.email,
        models.Candidates.name,
        models.Candidates.whatsapp,
        models.Candidates.pengalaman_total,  # total_experience
        models.Candidates.gelar_tertinggi,    # highest_degree
        models.Candidates.ekspektasi_gaji,    # salary_expectation
        models.Candidates.lokasi,             # domicile
        models.Candidates.skills,             # programming languages/skills
    ))
    # Date filters (commented out - applied_at doesn't exist in DB yet)
  # if start_date:
  #     try:
  #         start = datetime.strptime(start_date, "%Y-%m-%d")
  #     except ValueError:
  #         raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD.")
  #     query = query.where(models.Candidates.applied_at >= start.date())

  # if end_date:
  #     try:
  #         end = datetime.strptime(end_date, "%Y-%m-%d")
  #     except ValueError:
  #         raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD.")
  #     end = end + timedelta(days=1)  # inclusive end
  #     query = query.where(models.Candidates.applied_at <= end.date())

  # Commented out - processed_status doesn't exist in DB yet
  # if candidate_status:
  #     query = query.where(cast(models.Candidates.processed_status, String) == candidate_status.value)

  result = await db.execute(query)
  candidates = result.scalars().all()
  
  # Convert DB objects to dict with field mapping
  report_candidates = []
  for c in candidates:
      report_candidates.append({
          "email": c.email,
          "name": c.name,
          "whatsapp": c.whatsapp,
          "total_experience": float(c.pengalaman_total) if c.pengalaman_total else None,
          "highest_degree": c.gelar_tertinggi,
          "salary_expectation": None,  # ekspektasi_gaji is string, needs parsing
          "domicile": c.lokasi,
          "processed_status": "N/A",  # Field doesn't exist yet
          "primary_programming_language": c.skills.split(",")[0].strip() if c.skills else None,
          "programming_language_experience": c.skills,
      })
  
  if len(report_candidates) <= 0:
    raise HTTPException(status_code=400, detail="No candidate data found for the given filters.")
  # ===== Excel Export =====
  if file_type == "excel":
      wb = Workbook()
      ws = wb.active
      ws.title = "Candidates"

      headers = [
          "Email", "Name", "WhatsApp", "Total Exp (y)", "Primary Lang",
          "Lang Experience", "Domicile", "Status", "Salary (IDR)", "Highest Degree"
      ]
      ws.append(headers)

      # Header style
      header_font = Font(bold=True)
      header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
      for col_idx in range(1, len(headers) + 1):
          cell = ws.cell(row=1, column=col_idx)
          cell.font = header_font
          cell.fill = header_fill
          cell.alignment = Alignment(vertical="top")

      # Data rows (keep numbers as numbers when available)
      for r in report_candidates:
          ws.append([
              r["email"],
              r["name"] or "-",
              r["whatsapp"] or "-",
              r["total_experience"],
              r["primary_programming_language"] or "-",
              r["programming_language_experience"] or "-",
              r["domicile"] or "-",
              r["processed_status"] or "-",
              r["salary_expectation"],
              r["highest_degree"] or "-",
          ])

      # Number formats
      # Total Exp (col 4), Salary (col 9)
      for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
          for cell in row:
              if isinstance(cell.value, (int, float)):
                  cell.number_format = "0.0"  # one decimal year
      for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
          for cell in row:
              if isinstance(cell.value, (int, float)):
                  cell.number_format = "#,##0"  # Excel will localize display

      # Freeze header row
      ws.freeze_panes = "A2"

      # Auto-fit columns based on max string length
      for col_idx in range(1, ws.max_column + 1):
          col_letter = get_column_letter(col_idx)
          max_len = 0
          for cell in ws[col_letter]:
              val = cell.value
              length = len(str(val)) if val is not None else 0
              if length > max_len:
                  max_len = length
          # Add a little padding
          ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

      buf = BytesIO()
      wb.save(buf)
      xlsx_bytes = buf.getvalue()
      buf.close()

      return Response(
          xlsx_bytes,
          media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
          headers={"Content-Disposition": 'attachment; filename="candidates.xlsx"'}
      )
  # PDF Generator
  rows = [
        ["Email","Name","WhatsApp","Total Exp (y)","Primary Lang","Lang Experience",
         "Domicile","Status","Salary (IDR)","Highest Degree"],
    ]
  for r in report_candidates:
    rows.append([
        r["email"],
        r["name"] or "-",
        r["whatsapp"] or "-",
        f"{r['total_experience']}" if r["total_experience"] is not None else "-",
        r["primary_programming_language"] or "-",
        r["programming_language_experience"] or "-",
        r["domicile"] or "-",
        r["processed_status"] or "-",
        f"{r['salary_expectation']:,.0f}".replace(",", ".") if r["salary_expectation"] is not None else "-",
        r["highest_degree"] or "-",
    ])
  buf = BytesIO()
  doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=16, rightMargin=16, topMargin=16, bottomMargin=18)
  styles = getSampleStyleSheet()
  story = [Paragraph("Candidate Report", styles["Title"]), Spacer(1, 6)]

  table = LongTable(rows, repeatRows=1)
  table.setStyle(TableStyle([
      ("GRID", (0,0), (-1,-1), 0.4, colors.grey),
      ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
      ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
      ("FONTSIZE", (0,0), (-1,-1), 8),
      ("ALIGN", (3,1), (3,-1), "RIGHT"),  # experience
      ("ALIGN", (8,1), (8,-1), "RIGHT"),  # salary
      ("VALIGN", (0,0), (-1,-1), "TOP"),
  ]))
  story.append(table)
  doc.build(story)

  pdf_bytes = buf.getvalue()
  buf.close()
  return Response(pdf_bytes, 
                  media_type="application/pdf",
                  headers={"Content-Disposition": 'inline; filename="candidates.pdf'}
                  )